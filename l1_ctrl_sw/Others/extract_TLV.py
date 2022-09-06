from openpyxl import load_workbook

input_filename = r'nFAPI_Cplane_Uplane_Configuration_V2.xlsx'
output_filename = r'_mem_copy.txt'
num_examples = 1

def hex_val(d_type,d_string):
	if d_type[-1] == r']' :
		arr_size = int(d_type[-2])
		d_string = d_string.split(r',')
		encoded = ""

		if d_type[0:-4] == 'uint8_t' :
			for i in range(0,arr_size):
				h_encoded = format(int(d_string[i]), 'x')
				encoded = encoded + h_encoded.zfill(2)
		elif d_type[0:-4] == 'uint16_t' :
			for i in range(0,arr_size):
				h_encoded = format(int(d_string[i]), 'x')
				encoded = encoded + h_encoded.zfill(4)
		elif d_type[0:-4] == 'uint32_t' :
			for i in range(0,arr_size):
				h_encoded = format(int(d_string[i]), 'x')
				encoded = encoded + h_encoded.zfill(8)
		else :
			encoded = ''

	else :
		if d_type == 'uint8_t' :
			if d_string == '???' :
				encoded = '??'
			else :
				encoded = format(int(d_string), 'x')
				encoded = encoded.zfill(2)
		elif d_type == 'uint16_t' :
			if d_string == '???' :
				encoded = '????'
			else :
				encoded = format(int(d_string), 'x')
				encoded = encoded.zfill(4)
		elif d_type == 'uint32_t' :
			if d_string == '???' :
				encoded = '????????'
			else :
				encoded = format(int(d_string), 'x')
				encoded = encoded.zfill(8)
		else :
			encoded = ''

	return encoded

wb = load_workbook(input_filename, data_only=True)
for example in range(1,num_examples+1):
	ws = wb['Example '+str(example)+' PARAM message']
	op = open('Example_'+str(example)+output_filename,'w+')

	message_header = []
	message_header_TLV1 = []
	num_TLV1 = 0
	message_header_TLV2 = []
	num_TLV2 = 0
	message_header_TLV3 = []
	num_TLV3 = 0

	for row in range(1,1001):
		cell_whole = ws['A'+str(row)]
		if cell_whole.font.b == True :
			op.writelines(message_header)
			op.writelines(message_header_TLV1)
			op.writelines(message_header_TLV2)
			message_header = []
			message_header_TLV1 = []
			message_header_TLV2 = []

			L = ["\n######## " , cell_whole.value , " ########\n\n"]
			op.writelines(L)

		elif not (cell_whole.value is None) :
			d_type = ws['B'+str(row)].value
			d_string = ws['C'+str(row)].value
			val = hex_val(d_type,d_string)
			message_header.append(val)
			message_header.append("\n")

		else :
			tlv = ws['B'+str(row)]
			if tlv.value is None :
				tlv = ws['C'+str(row)]
				if tlv.value is None :
					tlv = ws['D'+str(row)]
					if tlv.value is None :
						continue

					elif tlv.value.lower() == 'tag' :
						num_TLV3 = num_TLV3 + 1

						d_string = ws['F'+str(row)].value
						val = d_string[2:].zfill(4)
						message_header_TLV3.append(val)
						message_header_TLV3.append("\n")

					elif tlv.value.lower() == 'length' :
						d_type = ws['E'+str(row)].value
						d_string = ws['F'+str(row)].value
						val = hex_val(d_type,d_string)
						message_header_TLV3.append(val)
						message_header_TLV3.append("\n")

					elif tlv.value.lower() == 'value' :
						d_type = ws['E'+str(row)].value
						d_string = ws['F'+str(row)].value
						val = hex_val(d_type,d_string)
						message_header_TLV3.append(val)
						message_header_TLV3.append("\n")

				elif tlv.value.lower() == 'tag' :
					num_TLV2 = num_TLV2 + 1

					d_string = ws['E'+str(row)].value
					val = d_string[2:].zfill(4)
					message_header_TLV2.append(val)
					message_header_TLV2.append("\n")

				elif tlv.value.lower() == 'length' :
					d_type = ws['D'+str(row)].value
					d_string = ws['E'+str(row)].value
					val = hex_val(d_type,d_string)
					message_header_TLV2.append(val)
					message_header_TLV2.append("\n")

				elif tlv.value.lower() == 'value' :
					d_type = ws['D'+str(row)].value
					d_string = ws['E'+str(row)].value
					val = hex_val(d_type,d_string)
					message_header_TLV2.append(val)
					message_header_TLV2.append("\n")

			elif tlv.value.lower() == 'tag' :
				num_TLV1 = num_TLV1 + 1

				d_string = ws['D'+str(row)].value
				val = d_string[2:].zfill(4)
				message_header_TLV1.append(val)
				message_header_TLV1.append("\n")

			elif tlv.value.lower() == 'length' :
				d_type = ws['C'+str(row)].value
				d_string = ws['D'+str(row)].value
				val = hex_val(d_type,d_string)
				message_header_TLV1.append(val)
				message_header_TLV1.append("\n")

			elif tlv.value.lower() == 'value' :
				d_type = ws['C'+str(row)].value
				d_string = ws['D'+str(row)].value
				val = hex_val(d_type,d_string)
				val = val.zfill(2*int(message_header_TLV1[-2],16))
				message_header_TLV1.append(val)
				message_header_TLV1.append("\n")

	op.writelines(message_header)
	op.writelines(message_header_TLV1)
	op.writelines(message_header_TLV2)

op.close()